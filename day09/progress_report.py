("""Parse `subjects.txt` and return open/closed assignments.

Usage:
  - Import functions: `load_subjects`, `open_assignments`, `closed_assignments`.
  - Run as script: `python3 progress_report.py [--json]`
""")

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List


def load_subjects(path: str | Path = "subjects.txt") -> List[Dict[str, str]]:
	"""Load and parse the subjects file.

	Treats columns as:
	  1) serial number (int)
	  2) progress status (OPEN/CLOSED)
	  3) assignment name (string)
	  4) submission date/time (ISO string)

	Returns a list of dicts with keys: `serial`, `status`, `assignment`, `submitted_at`.
	"""
	p = Path(path)
	records: List[Dict[str, str]] = []
	if not p.exists():
		raise FileNotFoundError(f"Subjects file not found: {p}")

	with p.open(encoding="utf-8") as fh:
		for raw in fh:
			line = raw.strip()
			if not line:
				continue
			# Split on tabs and drop empty fragments (some lines have extra tabs)
			parts = [part.strip() for part in line.split("\t") if part.strip() != ""]
			if len(parts) < 4:
				# If fewer than 4 parts, try splitting on multiple spaces as fallback
				parts = [part.strip() for part in line.split() if part.strip() != ""]
			if len(parts) < 4:
				# skip malformed line
				continue

			serial = parts[0]
			status = parts[1]
			assignment = parts[2]
			submitted_at = parts[3]

			records.append({
				"serial": serial,
				"status": status.upper(),
				"assignment": assignment,
				"submitted_at": submitted_at,
			})

	return records


def closed_assignments(records: List[Dict[str, str]]) -> List[Dict[str, str]]:
	"""Return records where status == 'CLOSED'."""
	return [r for r in records if r.get("status", "") == "CLOSED"]


def open_assignments(records: List[Dict[str, str]]) -> List[Dict[str, str]]:
	"""Return records where status != 'CLOSED'."""
	return [r for r in records if r.get("status", "") != "CLOSED"]


def _print_list(title: str, items: List[Dict[str, str]]):
	print(f"{title} ({len(items)})")
	for r in items:
		print(f"- {r['serial']}\t{r['assignment']}\t{r['submitted_at']}")


def _extract_days_and_final(assignment_text: str) -> tuple[set[int], bool]:
	"""From the assignment text, return a set of day numbers (1-8) and whether it's a final proposal.

	Matching is case-insensitive and ignores spaces/variations like 'Day05', 'day 05', 'Day 05 and 06'.
	"""
	import re

	text = assignment_text.lower()
	days = set()

	# find occurrences like 'day01', 'day 01', 'day1', 'day 1'
	for m in re.findall(r"day\s*0*([1-8])", text):
		try:
			days.add(int(m))
		except ValueError:
			pass

	# also catch patterns like 'day03andday04' if spaces removed
	for m in re.findall(r"day0*([1-8])", text.replace(" ", "")):
		try:
			days.add(int(m))
		except ValueError:
			pass

	is_final = "final" in text and ("proposal" in text or "project" in text)

	return days, is_final


def _extract_student_name(assignment_text: str) -> str:
	"""Extract the student name from an assignment string using heuristics.

	Heuristics (in order):
	  - text after the last ' by ' (case-insensitive)
	  - text after the last '-' character
	  - remove known assignment words and return the leftover
	"""
	import re

	text = assignment_text.strip()
	# try 'by' pattern
	m = re.search(r"\bby\b\s*(.+)$", text, flags=re.IGNORECASE)
	if m:
		return m.group(1).strip()

	# try dash
	if "-" in text:
		name = text.split("-")[-1].strip()
		if name:
			return name

	# fallback: remove assignment-related words
	cleaned = re.sub(r"(?i)day\s*0*[1-8]|day[s]?|final|project|proposal|and|for|the", "", text)
	# remove leftover punctuation
	cleaned = re.sub(r"[\-_:]+", " ", cleaned)
	cleaned = " ".join(cleaned.split())
	return cleaned.strip()


def submission_completeness_report(records: List[Dict[str, str]]) -> Dict[str, Dict[str, object]]:
	"""Return a report mapping student name -> {'days': set[int], 'final': bool, 'complete': bool}.

	Full submission is defined as having days 1 through 8 and a final project proposal.
	"""
	students: Dict[str, Dict[str, object]] = {}

	for r in records:
		assignment = r.get("assignment", "")
		days, is_final = _extract_days_and_final(assignment)
		name = _extract_student_name(assignment)
		if not name:
			# if we couldn't parse a name, skip
			continue

		entry = students.setdefault(name, {"days": set(), "final": False})
		entry["days"].update(days)
		if is_final:
			entry["final"] = True

	# compute completeness
	# Day07 had no assignment — ignore day 7
	required_days = {1, 2, 3, 4, 5, 6, 8}
	for name, info in students.items():
		info["complete"] = info["final"] and (required_days.issubset(info["days"]))

	return students


def _parse_iso_datetime(s: str) -> datetime:
	from datetime import datetime
	# input looks like: 2026-01-04T09:32:25Z
	try:
		return datetime.strptime(s, "%Y-%m-%dT%H:%M:%SZ")
	except Exception:
		# fallback: try without trailing Z
		return datetime.fromisoformat(s)


def export_report_xlsx(records: List[Dict[str, str]], filename: str = "report.xlsx"):
	"""Export two-sheet Excel workbook:
	  - Sheet 'Assignments': table of all open and closed assignments
	  - Sheet 'Students': one row per student and a column per assignment (days 01-06, 08, proposal)

	Cells are colored: on-time = black, late = yellow, missing = red.
	"""
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Font
	except Exception as exc:
		raise RuntimeError("openpyxl is required for export. Install with: pip install openpyxl") from exc

	from datetime import datetime

	# deadlines (naive datetimes)
	deadlines = {
		"day01": datetime.strptime("2025.11.01 22:00", "%Y.%m.%d %H:%M"),
		"day02": datetime.strptime("2025.11.09 22:00", "%Y.%m.%d %H:%M"),
		"day03": datetime.strptime("2025.11.16 22:00", "%Y.%m.%d %H:%M"),
		"day04": datetime.strptime("2025.11.23 22:00", "%Y.%m.%d %H:%M"),
		"day05": datetime.strptime("2025.11.29 22:00", "%Y.%m.%d %H:%M"),
		"day06": datetime.strptime("2025.12.06 22:00", "%Y.%m.%d %H:%M"),
		# note: day07 ignored
		"day08": datetime.strptime("2025.12.30 22:00", "%Y.%m.%d %H:%M"),
		"proposal": datetime.strptime("2026.01.11 22:00", "%Y.%m.%d %H:%M"),
	}

	# assignment keys order for columns
	assignment_keys = [f"day{d:02d}" for d in (1, 2, 3, 4, 5, 6)] + ["day08", "proposal"]

	# build per-student earliest submission datetimes for each assignment key
	students_map: Dict[str, Dict[str, datetime | None]] = {}

	for r in records:
		assignment = r.get("assignment", "")
		days, is_final = _extract_days_and_final(assignment)
		name = _extract_student_name(assignment)
		if not name:
			continue

		sub_dt = None
		try:
			sub_dt = _parse_iso_datetime(r.get("submitted_at", ""))
		except Exception:
			sub_dt = None

		entry = students_map.setdefault(name, {k: None for k in assignment_keys})

		for d in days:
			if d == 7:
				continue
			key = f"day{d:02d}"
			prev = entry.get(key)
			if sub_dt is None:
				continue
			if prev is None or sub_dt < prev:
				entry[key] = sub_dt

		if is_final:
			prev = entry.get("proposal")
			if sub_dt is not None and (prev is None or sub_dt < prev):
				entry["proposal"] = sub_dt

	# Create workbook
	wb = Workbook()
	ws1 = wb.active
	ws1.title = "Assignments"

	# Header
	ws1.append(["serial", "status", "assignment", "submitted_at"])
	for r in records:
		ws1.append([r.get("serial", ""), r.get("status", ""), r.get("assignment", ""), r.get("submitted_at", "")])

	# Students sheet
	ws2 = wb.create_sheet("Students")
	header = ["student"] + [k.upper() for k in assignment_keys]
	ws2.append(header)

	# Color definitions
	color_black = "000000"
	color_yellow = "FFFF00"
	color_red = "FF0000"

	for name, entry in sorted(students_map.items()):
		row = [name]
		for key in assignment_keys:
			dt = entry.get(key)
			if dt is None:
				row.append("Not submitted")
			else:
				# compare to deadline
				dl = deadlines.get(key)
				if dl is None:
					row.append("Submitted")
				else:
					row.append(dt.isoformat(sep=" ", timespec="seconds"))
		ws2.append(row)

	# Apply colors per cell in Students sheet
	for r_idx, row in enumerate(ws2.iter_rows(min_row=2, max_col=len(header), max_row=ws2.max_row), start=2):
		# r_idx matches Excel row index
		for c_idx, cell in enumerate(row[1:], start=2):  # skip student name column
			val = cell.value
			col_key = assignment_keys[c_idx - 2]
			if val == "Not submitted":
				cell.font = Font(color=color_red)
			else:
				# parse the string back to datetime to compare
				try:
					parsed = datetime.fromisoformat(val.replace(" ", "T"))
					dl = deadlines.get(col_key)
					if dl is not None and parsed > dl:
						cell.font = Font(color=color_yellow)
					else:
						cell.font = Font(color=color_black)
				except Exception:
					cell.font = Font(color=color_black)

	wb.save(filename)



def print_missing_report(records: List[Dict[str, str]]):
	report = submission_completeness_report(records)
	missing = {name: info for name, info in report.items() if not info.get("complete")}

	if not missing:
		print("All students have full submissions (Day01-08 + final project proposal).")
		return

	print(f"Students missing submissions: {len(missing)}")
	for name, info in sorted(missing.items()):
		days_done = sorted(info["days"]) if "days" in info else []
		final = info.get("final", False)
		missing_days = sorted(list(set(range(1, 9)) - set(days_done)))
		print(f"- {name}: days submitted={days_done}, final_submitted={final}, missing_days={missing_days}")


def main():
	parser = argparse.ArgumentParser(description="Print open and closed assignments from subjects.txt")
	parser.add_argument("--json", action="store_true", help="Output JSON with keys 'open' and 'closed'")
	parser.add_argument("--file", default="subjects.txt", help="Path to subjects file")
	parser.add_argument("--report", action="store_true", help="Print report of students missing submissions (Day01-08 + final)")
	parser.add_argument("--export", nargs="?", const="report.xlsx", help="Export an Excel workbook (.xlsx) with the assignments and students sheets. Optionally pass filename.")
	args = parser.parse_args()

	records = load_subjects(args.file)
	closed = closed_assignments(records)
	open_ = open_assignments(records)

	if args.json:
		out = {"open": open_, "closed": closed}
		print(json.dumps(out, ensure_ascii=False, indent=2))
	else:
		_print_list("Closed assignments", closed)
		print()
		_print_list("Open assignments", open_)

	# optional report: missing submissions
	if args.report:
		print()
		print_missing_report(records)

	if args.export:
		fname = args.export if isinstance(args.export, str) else "report.xlsx"
		print(f"Exporting workbook to {fname}...")
		export_report_xlsx(records, fname)
		print("Export complete.")


if __name__ == "__main__":
	main()

